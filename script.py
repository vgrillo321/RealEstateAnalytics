from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import urllib
import json

# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter

filename = '.\Estimados-Potencial-Renta.xlsx'

# Access the spreadsheet to be updated
wb = load_workbook(filename)

ws1 = wb.active
ws2 = wb.copy_worksheet(ws1) # Copy template to new sheet

state = input("Please add the state acronym (ex. FL)")
cityinput= input("Please add the city (optional)")
city = ""

# Create query string; create a function that can be applied to other values
for i in range(len(cityinput)):
    if cityinput[i] == " ":
        city+="%20"
        continue
    city += cityinput[i]
    
#Query string for the api call
url_q = "city="+city+"&state="+state

url = "https://api.rentcast.io/v1/listings/rental/long-term?"+url_q

headers = {
    "accept": "application/json",
    "X-Api-Key": "4bbec6c728f54cfd9e408d03c65e9870" # TODO: Pass API key safer
}

# Call API
response = requests.get(url, headers=headers)

# Add response as text data
data = response.text

#Function to pass the payload data needed.
def passDataExcel(address,price,size):

    address_cell = 'B'+ str(row_counter)
    price_cell = 'C'+ str(row_counter)
    size_cell = 'D'+ str(row_counter)

    # Add values on each cell
    ws2[address_cell] = address
    ws2[price_cell] = price
    ws2[size_cell] = size

    wb.save(filename)


# Load json data into a variable
with open("sample.json", "w") as openfile:
    openfile.write(data)
    
with open("sample.json", "r") as openfile:
    json_object = json.load(openfile)

#Pass Json data to spreadsheet if it contains the squareFootage
row_counter = 9
jsonObjectNumber = 0

while row_counter < 19:

    address = json_object[jsonObjectNumber]["addressLine1"]
    price = json_object[jsonObjectNumber]["price"]
    
    if not (json_object[jsonObjectNumber].get('squareFootage') is None):
        sqft = json_object[jsonObjectNumber]["squareFootage"]
        passDataExcel(address, price, sqft)
        row_counter+=1
    
    jsonObjectNumber+=1





